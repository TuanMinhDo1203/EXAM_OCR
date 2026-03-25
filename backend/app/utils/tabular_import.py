from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_value(value: object) -> str:
    return str(value or "").strip()


def load_tabular_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        decoded = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        if reader.fieldnames is None:
            return []
        normalized_headers = [_normalize_header(header) for header in reader.fieldnames]
        rows: list[dict[str, str]] = []
        for row in reader:
            item = {
                normalized_headers[index]: _normalize_value(value)
                for index, value in enumerate(row.values())
            }
            if any(item.values()):
                rows.append(item)
        return rows

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration:
            return []
        normalized_headers = [_normalize_header(header) for header in headers]
        rows: list[dict[str, str]] = []
        for values in iterator:
            item = {
                normalized_headers[index]: _normalize_value(value)
                for index, value in enumerate(values)
                if index < len(normalized_headers)
            }
            if any(item.values()):
                rows.append(item)
        return rows

    raise ValueError("Unsupported file type. Please upload a CSV or XLSX file.")
